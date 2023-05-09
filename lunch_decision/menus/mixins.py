from openpyxl.reader.excel import load_workbook


def read_excel_file(file):
    data = []
    wb = load_workbook(file)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, values_only=True):
        item_name, description, price, image_url, is_active = row
        data.append({
            'name': item_name,
            'description': description,
            'price': price,
            'image': image_url,
            'is_active': is_active,
        })

    return data
